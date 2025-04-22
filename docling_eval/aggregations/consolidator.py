import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from pandas import DataFrame

from docling_eval.aggregations.multi_evalutor import MultiEvaluation
from docling_eval.datamodels.types import ConsolidationFormats, EvaluationModality
from docling_eval.evaluators.base_evaluator import EvaluationRejectionType
from docling_eval.evaluators.bbox_text_evaluator import DatasetBoxesTextEvaluation
from docling_eval.evaluators.layout_evaluator import DatasetLayoutEvaluation
from docling_eval.evaluators.markdown_text_evaluator import DatasetMarkdownEvaluation
from docling_eval.evaluators.readingorder_evaluator import DatasetReadingOrderEvaluation
from docling_eval.evaluators.stats import DatasetStatistics
from docling_eval.evaluators.table_evaluator import DatasetTableEvaluation

_log = logging.getLogger(__name__)


def export_value(val: Union[float, DatasetStatistics]) -> str:
    r"""Get statistics value"""
    if isinstance(val, DatasetStatistics):
        fmt_val = f"{val.mean:.2f}±{val.std:.2f}"
    else:
        fmt_val = f"{val:.2f}"

    return fmt_val


class Consolidator:
    r"""
    Consolidate a MultiEvaluation into a comparison matrix

    The comparison matrix has 3 dimensions:
    - Benchmarks
    - ConversionProviders
    - Modalities
    """

    def __init__(self, output_path: Path):
        r""" """
        self._output_path = output_path
        self._excel_engine = "openpyxl"
        self._sheet_name = "matrix"
        self._excel_filename = "consolidation_matrix.xlsx"

        self._output_path.mkdir(parents=True, exist_ok=True)

    def __call__(
        self,
        multi_evaluation: MultiEvaluation,
        consolidation_format: Optional[
            ConsolidationFormats
        ] = ConsolidationFormats.EXCEL,
    ) -> Tuple[Dict[EvaluationModality, DataFrame], Optional[Path]]:
        r""" """
        dfs = self._build_dataframes(multi_evaluation)

        # Export dataframe
        if consolidation_format == ConsolidationFormats.EXCEL:
            produced_fn = self._to_excel(dfs)
            _log.info("Produced excel: %s", str(produced_fn))
        else:
            _log.info("Unsupported consolidation format: %s", consolidation_format)

        return dfs, produced_fn

    def _to_excel(self, dfs: Dict[EvaluationModality, DataFrame]) -> Path:
        r""" """
        excel_fn = self._output_path / self._excel_filename
        startrow = 0
        header_rows: List[int] = []
        with pd.ExcelWriter(excel_fn, engine=self._excel_engine) as writer:  # type: ignore
            for modality, df in dfs.items():
                if self._sheet_name in writer.book.sheetnames:
                    sheet = writer.book[self._sheet_name]
                    startrow = sheet.max_row + 2

                # Add the modality as a "header" for the metrics subtable
                header_df = DataFrame([modality.name])
                header_rows.append(startrow + 1)
                header_df.to_excel(
                    writer,
                    sheet_name=self._sheet_name,
                    startrow=startrow,
                    index=False,
                    header=False,
                )
                startrow += 1

                # Metrics subtable
                df.to_excel(
                    writer,
                    sheet_name=self._sheet_name,
                    startrow=startrow,
                    index=False,
                )
        # Format the excel
        self._format_excel(excel_fn, header_rows)

        return excel_fn

    def _format_excel(self, excel_fn: Path, header_rows: List[int]):
        r"""Do some proper formatting of the generated excel"""
        workbook = load_workbook(excel_fn)
        sheet = workbook[self._sheet_name]

        # Adjust the cell width
        for col in sheet.columns:
            # Find the maximum length of strings in this column (excluding empty cells)
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2  # Add some padding to make it look better
            first_cell = col[0]
            assert isinstance(first_cell, Cell)
            sheet.column_dimensions[first_cell.column_letter].width = adjusted_width

        # Iterate through each cell in the worksheet and remove borders
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = openpyxl.styles.Border()  # Remove borders

        # Make bold the subtable headers
        bold_font = Font(bold=True)
        for header_row in header_rows:
            cell = sheet.cell(row=header_row, column=1)
            cell.font = bold_font
            x = 0

        # Save back the excel
        workbook.save(excel_fn)

    def _build_dataframes(
        self,
        multi_evaluation: MultiEvaluation,
    ) -> Dict[EvaluationModality, DataFrame]:
        r"""
        Return a Dict with dataframes per modality
        """
        # Collect all data to build the dataframes
        df_data: Dict[EvaluationModality, List[Dict[str, Union[str, float, int]]]] = {}

        # Collect the dataframe data
        for benchmark, prov_mod_eval in multi_evaluation.evaluations.items():
            for experiment, mod_eval in prov_mod_eval.items():
                for modality, single_evaluation in mod_eval.items():
                    evaluation = single_evaluation.evaluation

                    if modality == EvaluationModality.LAYOUT:
                        metrics = self._layout_metrics(evaluation)
                    elif modality == EvaluationModality.MARKDOWN_TEXT:
                        metrics = self._markdowntext_metrics(evaluation)
                    elif modality == EvaluationModality.TABLE_STRUCTURE:
                        metrics = self._tablestructure_metrics(evaluation)
                    elif modality == EvaluationModality.READING_ORDER:
                        metrics = self._readingorder_metrics(evaluation)
                    elif modality == EvaluationModality.BBOXES_TEXT:
                        metrics = self._bboxestext_metrics(evaluation)
                    else:
                        _log.error(
                            "Evaluation modality unsupported for export: %s", modality
                        )
                        continue

                    # Gather the dataframe data
                    provider = (
                        single_evaluation.prediction_provider_type.value
                        if single_evaluation.prediction_provider_type is not None
                        else "Unkown"
                    )
                    data: Dict[str, Union[str, float]] = {
                        "Benchmark": benchmark.value,
                        "Provider": provider,
                        "Experiment": experiment,
                        "evaluated_samples": evaluation.evaluated_samples,
                    }
                    for rej_type in EvaluationRejectionType:
                        if rej_type not in evaluation.rejected_samples:
                            data[rej_type.value] = 0
                        else:
                            data[rej_type.value] = evaluation.rejected_samples[rej_type]

                    data |= metrics
                    if modality not in df_data:
                        df_data[modality] = []
                    df_data[modality].append(data)

        # Build the dataframes
        dfs: Dict[EvaluationModality, DataFrame] = {}
        for modality, m_data in df_data.items():
            df = DataFrame(m_data)
            df = df.sort_values(by=["Benchmark", "Provider"], ascending=[True, True])
            dfs[modality] = df

        return dfs

    def _layout_metrics(self, evaluation: DatasetLayoutEvaluation) -> Dict[str, str]:
        r"""Get the metrics for the LayoutEvaluation"""
        metrics = {
            "mAP": export_value(evaluation.map_stats),
            "mAP_50": export_value(evaluation.map_50_stats),
            "mAP_75": export_value(evaluation.map_75_stats),
            "weighted_mAP_50": export_value(evaluation.weighted_map_50_stats),
            "weighted_mAP_75": export_value(evaluation.weighted_map_75_stats),
            "weighted_mAP_90": export_value(evaluation.weighted_map_90_stats),
            "weighted_mAP_95": export_value(evaluation.weighted_map_95_stats),
        }
        for class_evaluation in evaluation.evaluations_per_class:
            key = f"class_{class_evaluation.label}"
            metrics[key] = export_value(class_evaluation.value)

        return metrics

    def _markdowntext_metrics(
        self,
        evaluation: DatasetMarkdownEvaluation,
    ) -> Dict[str, str]:
        r""" """
        metrics = {
            "BLEU": export_value(evaluation.bleu_stats),
            "F1": export_value(evaluation.f1_score_stats),
            "Precision": export_value(evaluation.precision_stats),
            "Recall": export_value(evaluation.recall_stats),
            "Edit_Distance": export_value(evaluation.edit_distance_stats),
            "METEOR": export_value(evaluation.meteor_stats),
        }
        return metrics

    def _tablestructure_metrics(
        self,
        evaluation: DatasetTableEvaluation,
    ) -> Dict[str, str]:
        r""" """
        metrics = {
            "TEDS": export_value(evaluation.TEDS),
            "TEDS_struct": export_value(evaluation.TEDS_struct),
            "TEDS_simple": export_value(evaluation.TEDS_simple),
            "TEDS_complex": export_value(evaluation.TEDS_complex),
        }
        return metrics

    def _readingorder_metrics(
        self,
        evaluation: DatasetReadingOrderEvaluation,
    ) -> Dict[str, str]:
        r""" """
        metrics = {
            "ARD": export_value(evaluation.ard_stats),
            "Weighted_ARD": export_value(evaluation.w_ard_stats),
        }
        return metrics

    def _bboxestext_metrics(
        self,
        evaluation: DatasetBoxesTextEvaluation,
    ) -> Dict[str, str]:
        r""" """
        metrics = {
            "BLEU": export_value(evaluation.bleu_stats),
            "F1": export_value(evaluation.f1_score_stats),
            "Precision": export_value(evaluation.precision_stats),
            "Recall": export_value(evaluation.recall_stats),
            "Edit_Distance": export_value(evaluation.edit_distance_stats),
            "METEOR": export_value(evaluation.meteor_stats),
        }
        return metrics
